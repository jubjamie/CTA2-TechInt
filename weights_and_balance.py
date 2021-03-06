import numpy as np
from matplotlib import pyplot as plt
from matplotlib.ticker import FuncFormatter
from openpyxl import load_workbook
import filepath1

# Excel Sheet Loading
cg_file_path = filepath1.cg_file_path
cg_file = load_workbook(filename=cg_file_path, data_only=True)
cg_params = cg_file['CG']  # Load into CG Sheet

# Init list holders
param_name = []
param_100_w = []
param_80_w = []
param_100_x = []
param_80_x = []
param_100_m = []
param_80_m = []

# Loop helpers
data_available = True
data_row_start = 3  # Row where data starts

# Loop through variable names and values, saving to list.
while data_available is True:
    # print(cad_params["A"+str(data_row_start)].value)
    if cg_params["A" + str(data_row_start)].value is not None:
        param_name.append(cg_params["A" + str(data_row_start)].value)
        param_100_w.append(cg_params["B" + str(data_row_start)].value)
        param_80_w.append(cg_params["C" + str(data_row_start)].value)
        param_100_x.append(cg_params["F" + str(data_row_start)].value)
        param_80_x.append(cg_params["G" + str(data_row_start)].value)
        param_100_m.append(cg_params["H" + str(data_row_start)].value)
        param_80_m.append(cg_params["I" + str(data_row_start)].value)
        data_row_start = data_row_start+1
    else:
        break

# Zip into a dictionary
w_100 = dict(zip(param_name, param_100_w))
m_100 = dict(zip(param_name, param_100_m))
x_100 = dict(zip(param_name, param_100_x))
w_80 = dict(zip(param_name, param_80_w))
m_80 = dict(zip(param_name, param_80_m))
x_80 = dict(zip(param_name, param_80_x))

# Add statics
big_weights = {"MTOW_w": cg_params["Q4"].value,
               "MTOW_m": cg_params["Q3"].value,
               "MTOW_x": cg_params["Q5"].value,
               "OWE_w": cg_params["Q23"].value,
               "OWE_m": cg_params["Q22"].value,
               "OWE_x": cg_params["Q24"].value,
               "MZFW": cg_file["Masses"]["E28"].value
               }

ax_lims = [big_weights["OWE_w"]-1000, big_weights["MTOW_w"]*1.02]
c_bar = cg_params["M7"].value
h0 = cg_params["M8"].value
print(h0)
wing_pos = h0 * c_bar
seat_start_fwd = 6.304  # m
seat_start_aft = 21.290  # m
seat_pitch = 31 * 2.54 / 100

cad_file_path = filepath1.cad_file_path
cad_file = load_workbook(filename=cad_file_path, data_only=True)
cad_params = cad_file['Hold Dims']  # Load into CG Sheet
# Luggage
# Init list holders
param_name = []
param_value = []

# Loop helpers
data_available = True
data_row_start = 1  # Row where data starts

# Loop through variable names and values, saving to list.
while data_available is True:
    # print(cad_params["A"+str(data_row_start)].value)
    if cad_params["B"+str(data_row_start)].value is not None:
        param_name.append(cad_params["B"+str(data_row_start)].value)
        param_value.append(cad_params["C"+str(data_row_start)].value)
        data_row_start = data_row_start+1
    else:
        break

# Zip into a dictionary
hold_params = dict(zip(param_name, param_value))
#print(hold_params)

fuel_params = cg_file['Fuel']  # Load into Fuel Tank Sheet
# Luggage
# Init list holders
tank_vol = []
tank_arm = []
fuel_density = fuel_params['F22'].value
for tank in range(2, 19, 1):
    tank_vol.append(fuel_params['F'+str(tank)].value)
    tank_arm.append(fuel_params['B'+str(tank)].value)

#print(tank_arm)
#print(tank_vol)


def mac_x_point(mac_pos, w):
    return (mac_pos - 0.25) * w


def mac_axes(mac_pos):
    lower_ax_point = mac_x_point(mac_pos, ax_lims[0])
    upper_ax_point = mac_x_point(mac_pos, ax_lims[1])
    return [lower_ax_point, upper_ax_point]


def mom_calc(w, x):
    #print(x)
    #print(h0)
   # print(x/c_bar)
   # print(((x/c_bar) - h0))
    return w * ((x/c_bar) - h0)


def seat_loading_wa(seats=2, load_dir="fwd"):

    distances = []
    moms = []
    big_dist = []
    pax_mass = (95-23) * seats  # For both window seats
    for row in np.arange(1, 21, 1):
        local_dist = ((seat_start_fwd + ((row - 1) * seat_pitch))/c_bar) - h0
        distances.append(local_dist)
        big_dist.append(seat_start_fwd + ((row - 1) * seat_pitch))
        moms.append(pax_mass * local_dist)

    if load_dir.lower() == "aft":
        distances.reverse()
        moms.reverse()

    return {"distances": distances, "moments": moms, "pax mass": pax_mass}


def fuel_loading():
    distances = []
    moms = []
    tank_masses = []
    for tank_id, this_tank_vol in enumerate(tank_vol):
        local_dist = (tank_pos(tank_arm[tank_id])/c_bar) - h0
        distances.append(local_dist)
        tank_mass = fuel_density * this_tank_vol * 2
        tank_masses.append(tank_mass)
        moms.append(local_dist * tank_mass)
    distances.reverse()
    moms.reverse()
    tank_masses.reverse()

    return{"distances": distances, "moments": moms, "tank mass": tank_masses}


def tank_pos(ob_dist):
    tank_centre_c_pos = 0.4
    tank_centre_sweep = np.arctan((cad_file['Interface']['B101'].value*(0.5-tank_centre_c_pos))/(cad_file['Interface']['B98'].value/2))  # 0.1*cr/halfspan
    print("tank sweep: " + str(tank_centre_sweep))
    offset = np.tan(tank_centre_sweep) * ob_dist
    total_pos = offset + cad_file['Sheet2']['R3'].value
    return total_pos

def noseWheel():
    nosepos = cad_file['Sheet1']['B30'].value
    mainpos = cad_file['Sheet1']['B34'].value
    condition_point = (0.975*(mainpos-nosepos)) + nosepos
    return condition_point/c_bar


def to_tons(x, pos):
    'The two args are the value and tick position'
    return '%1.1fT' % (x*1e-3)


def to_mac(x, pos):
    valf = int(round(100*(x/ax_lims[0]+0.25)))
    if valf >= 0:
        return str(valf) + '%'
    else:
        return " "


formattery = FuncFormatter(to_tons)
formatterx = FuncFormatter(to_mac)


def plotit(mac_range=[0.11, 0.51]):
    fig = plt.figure(figsize=(16, 9))
    ax = fig.add_subplot(1, 1, 1)
    ax.yaxis.set_major_formatter(formattery)
    ax.xaxis.set_major_formatter(formatterx)
    ax.xaxis.set_ticks_position('bottom')
    ax.yaxis.grid(which="major", color='k', linestyle='-', linewidth=1)



    break_points={}
    """ This section plots static points"""
    # OWE point from CG Excel
    curr_weight = big_weights["OWE_w"]
    owe_mom = mom_calc(big_weights["OWE_w"], big_weights['OWE_x'])
    plt.plot(owe_mom, curr_weight, 'ro', markersize=10)
    bbox_props = dict(boxstyle="round", fc="w", ec="0.5", alpha=0.9)
    ax.text(owe_mom, curr_weight-500, "OWE", ha="center", va="center", size=11,
            bbox=bbox_props)
    break_points['OWE Weight'] = curr_weight

    # Plot target pax end point
    # plt.plot(mom_calc((curr_weight+7200), cg_params["Q50"].value), (curr_weight+7200), 'ro', markersize=10)

    # Plot MZFW and MTOW
    plt.plot([mac_axes(mac_range[0]-0.1)[0], mac_axes(mac_range[1]+0.05)[0]], [big_weights["MTOW_w"], big_weights["MTOW_w"]], 'r', lw=5)
    plt.plot([mac_axes(mac_range[0]-0.1)[0], mac_axes(mac_range[1]+0.05)[0]], [big_weights["MZFW"], big_weights["MZFW"]], 'r--')

    """ This section plots the loops front first"""
    curr_mom = owe_mom
    seat_fwd = seat_loading_wa(2)
    seat_window_loop_moms = [curr_mom]
    seat_window_loop_weights = [curr_weight]

    # Window Seats
    init_mom = curr_mom
    for row_id, row in enumerate(seat_fwd["moments"]):
        curr_weight = curr_weight + seat_fwd["pax mass"]
        curr_mom = curr_mom + row
        seat_window_loop_moms.append(curr_mom)
        seat_window_loop_weights.append(curr_weight)

    break_points['Window Weight'] = curr_weight
    # Draw loop label
    text_y = np.mean([break_points["OWE Weight"], curr_weight])
    text_x = np.mean([curr_mom, init_mom])
    bbox_props = dict(boxstyle="round", fc="w", ec="0.5", alpha=0.9)
    ax.text(text_x, text_y, "Windows", ha="center", va="center", size=15,
            bbox=bbox_props)

    # Aisle Seats
    init_mom = curr_mom
    for row_id, row in enumerate(seat_fwd["moments"]):
        curr_weight = curr_weight + seat_fwd["pax mass"]
        curr_mom = curr_mom + row
        seat_window_loop_moms.append(curr_mom)
        seat_window_loop_weights.append(curr_weight)

    break_points['Aisle Weight'] = curr_weight
    # Draw loop label
    text_y = np.mean([break_points["Window Weight"], curr_weight])
    text_x = np.mean([curr_mom, init_mom])
    bbox_props = dict(boxstyle="round", fc="w", ec="0.5", alpha=0.9)
    ax.text(text_x, text_y, "Aisle", ha="center", va="center", size=15,
            bbox=bbox_props)

    seat_fwd = seat_loading_wa(1)
    init_mom = curr_mom
    for row_id, row in enumerate(seat_fwd["moments"]):
        curr_weight = curr_weight + seat_fwd["pax mass"]
        curr_mom = curr_mom + row
        seat_window_loop_moms.append(curr_mom)
        seat_window_loop_weights.append(curr_weight)

    break_points['Middle Weight'] = curr_weight
    # Draw loop label
    text_y = np.mean([break_points["Aisle Weight"], curr_weight])
    text_x = np.mean([curr_mom, init_mom])
    bbox_props = dict(boxstyle="round", fc="w", ec="0.5", alpha=0.9)
    ax.text(text_x, text_y, "Middle", ha="center", va="center", size=11,
            bbox=bbox_props)

    # Plot Window Loop
    plt.plot(seat_window_loop_moms, seat_window_loop_weights, 'b')

    init_mom = curr_mom
    # Plot Forward Hold Loop
    # Forward Case - Treat as point to point mass
    hold_fwd_case_loop_moms = [curr_mom]
    hold_fwd_case_loop_weights = [curr_weight]
    # Fwd hold
    hold_fwd_centre = np.mean([hold_params["hold_fwd_fwd"], hold_params["hold_fwd_aft"]])
    hold_fwd_centre_local = (hold_fwd_centre/c_bar) - h0
    hold_fwd_mom = hold_fwd_centre_local * hold_params["hold_fwd_case_fwd"]
    curr_mom = curr_mom + hold_fwd_mom
    curr_weight = curr_weight + hold_params["hold_fwd_case_fwd"]
    hold_fwd_case_loop_moms.append(curr_mom)
    hold_fwd_case_loop_weights.append(curr_weight)
    break_points["Hold FC Fwd"] = curr_weight
    # Aft
    hold_aft_centre = np.mean([hold_params["hold_aft_fwd"], hold_params["hold_aft_aft"]])
    hold_aft_centre_local = (hold_aft_centre / c_bar) - h0
    hold_aft_mom = hold_aft_centre_local * hold_params["hold_fwd_case_aft"]
    curr_mom = curr_mom + hold_aft_mom
    curr_weight = curr_weight + hold_params["hold_fwd_case_aft"]
    hold_fwd_case_loop_moms.append(curr_mom)
    hold_fwd_case_loop_weights.append(curr_weight)
    break_points["Hold FC Aft"] = curr_weight

    text_y = break_points["Hold FC Fwd"]
    text_x = curr_mom - 200
    bbox_props = dict(boxstyle="round", fc="w", ec="0.5", alpha=0.9)
    ax.text(text_x, text_y, "Baggage FC", ha="center", va="center", size=11,
            bbox=bbox_props)

    plt.plot(hold_fwd_case_loop_moms, hold_fwd_case_loop_weights, 'c')

    fuel_loops_weights = [curr_weight]
    fuel_loops_moms = [curr_mom]

    fuel_loader = fuel_loading()
    for tank_id, tank_mom in enumerate(fuel_loader['moments']):
        curr_mom = curr_mom + tank_mom
        fuel_loops_moms.append(curr_mom)
        curr_weight = curr_weight + fuel_loader['tank mass'][tank_id]
        fuel_loops_weights.append(curr_weight)
    break_points['Fuel End'] = curr_weight

    plt.plot(fuel_loops_moms, fuel_loops_weights, 'g')

    text_y_f = np.mean([break_points["Hold FC Aft"], break_points["Fuel End"]])
    text_x = curr_mom - 700
    bbox_props = dict(boxstyle="round", fc="w", ec="0.5", alpha=0.9)
    ax.text(text_x, text_y_f, "Fuel", ha="center", va="center", size=11,
            bbox=bbox_props)

    # Plot Aft Hold Loop
    # Aft Case - Treat as point to point mass
    curr_weight = break_points["Middle Weight"]
    curr_mom = init_mom
    hold_aft_case_loop_moms = [curr_mom]
    hold_aft_case_loop_weights = [curr_weight]
    # Aft hold
    hold_aft_centre = np.mean([hold_params["hold_aft_fwd"], hold_params["hold_aft_aft"]])
    hold_aft_centre_local = (hold_aft_centre / c_bar) - h0
    hold_aft_mom = hold_aft_centre_local * hold_params["hold_aft_case_aft"]
    curr_mom = curr_mom + hold_aft_mom
    curr_weight = curr_weight + hold_params["hold_aft_case_aft"]
    hold_aft_case_loop_moms.append(curr_mom)
    hold_aft_case_loop_weights.append(curr_weight)
    break_points["Hold AC Aft"] = curr_weight
    #Fwd
    hold_fwd_centre = np.mean([hold_params["hold_fwd_fwd"], hold_params["hold_fwd_aft"]])
    hold_fwd_centre_local = (hold_fwd_centre / c_bar) - h0
    hold_fwd_mom = hold_fwd_centre_local * hold_params["hold_aft_case_fwd"]
    curr_mom = curr_mom + hold_fwd_mom
    curr_weight = curr_weight + hold_params["hold_aft_case_fwd"]
    hold_aft_case_loop_moms.append(curr_mom)
    hold_aft_case_loop_weights.append(curr_weight)
    break_points["Hold AC Fwd"] = curr_weight

    text_y = break_points["Hold AC Aft"]
    text_x = curr_mom + 200
    bbox_props = dict(boxstyle="round", fc="w", ec="0.5", alpha=0.9)
    ax.text(text_x, text_y, "Baggage AC", ha="center", va="center", size=11,
            bbox=bbox_props)

    plt.plot(hold_aft_case_loop_moms, hold_aft_case_loop_weights, 'c')

    fuel_loops_weights = [curr_weight]
    fuel_loops_moms = [curr_mom]

    fuel_loader = fuel_loading()
    for tank_id, tank_mom in enumerate(fuel_loader['moments']):
        curr_mom = curr_mom + tank_mom
        fuel_loops_moms.append(curr_mom)
        curr_weight = curr_weight + fuel_loader['tank mass'][tank_id]
        fuel_loops_weights.append(curr_weight)
    break_points['Fuel End'] = curr_weight

    plt.plot(fuel_loops_moms, fuel_loops_weights, 'g')

    #get fuel loop basic info for envelopes
    fuel_x_delta = max(fuel_loops_moms)-min(fuel_loops_moms)

    #text_y = np.mean([break_points["Hold AC Aft"], break_points["Fuel End"]])
    text_x = curr_mom + 250
    bbox_props = dict(boxstyle="round", fc="w", ec="0.5", alpha=0.9)
    ax.text(text_x, text_y_f, "Fuel", ha="center", va="center", size=11,
            bbox=bbox_props)

    """
    # Plot Full Loop
    curr_weight = big_weights["OWE_w"]
    curr_mom = owe_mom
    seat_aft = seat_loading_wa(5, load_dir="aft")
    seat_window_loop_moms = [curr_mom]
    seat_window_loop_weights = [curr_weight]
    for row_id, row in enumerate(seat_aft["moments"]):
        curr_weight = curr_weight + seat_aft["pax mass"]
        curr_mom = curr_mom + row
        seat_window_loop_moms.append(curr_mom)
        seat_window_loop_weights.append(curr_weight)
    plt.plot(seat_window_loop_moms, seat_window_loop_weights, 'b--')
    # Plot Full Loop
    curr_weight = big_weights["OWE_w"]
    curr_mom = owe_mom
    seat_aft = seat_loading_wa(5, load_dir="fwd")
    seat_window_loop_moms = [curr_mom]
    seat_window_loop_weights = [curr_weight]
    for row_id, row in enumerate(seat_aft["moments"]):
        curr_weight = curr_weight + seat_aft["pax mass"]
        curr_mom = curr_mom + row
        seat_window_loop_moms.append(curr_mom)
        seat_window_loop_weights.append(curr_weight)
    plt.plot(seat_window_loop_moms, seat_window_loop_weights, 'b--')
    
    """

    """ Plot the rear loops"""
    curr_weight = big_weights["OWE_w"]
    curr_mom = owe_mom
    seat_aft = seat_loading_wa(2, load_dir="aft")
    seat_window_loop_moms = [curr_mom]
    seat_window_loop_weights = [curr_weight]
    for row_id, row in enumerate(seat_aft["moments"]):
        curr_weight = curr_weight + seat_aft["pax mass"]
        curr_mom = curr_mom + row
        seat_window_loop_moms.append(curr_mom)
        seat_window_loop_weights.append(curr_weight)

    for row_id, row in enumerate(seat_aft["moments"]):
        curr_weight = curr_weight + seat_aft["pax mass"]
        curr_mom = curr_mom + row
        seat_window_loop_moms.append(curr_mom)
        seat_window_loop_weights.append(curr_weight)

    seat_aft = seat_loading_wa(1, load_dir="aft")
    for row_id, row in enumerate(seat_aft["moments"]):
        curr_weight = curr_weight + seat_aft["pax mass"]
        curr_mom = curr_mom + row
        seat_window_loop_moms.append(curr_mom)
        seat_window_loop_weights.append(curr_weight)

    # Plot Window Loop
    plt.plot(seat_window_loop_moms, seat_window_loop_weights, 'b')

    print(break_points)
    plt.ylabel("Mass", size=14)
    plt.xlabel("% MAC", size=14)
    plt.xlim(mac_axes(mac_range[0]-0.01)[0], mac_axes(mac_range[1]-0.1)[0])
    ax_lims[1] = break_points["Fuel End"] * 1.02
    plt.ylim(ax_lims[0], ax_lims[1])

    """ Plot MAC axes """

    mac_spacing = 0.02
    mac_set = np.arange(mac_range[0], mac_range[1], mac_spacing)

    """Plot SC limits"""
    plt.plot(mac_axes((noseWheel()-h0)+0.25), ax_lims, 'r', zorder=10, lw=5)  # Nose Wheel Reaction
    ax.text(mac_axes((noseWheel()-h0)+0.25)[0] + 250, big_weights['OWE_w'] + 1400, "Ground Handling Envelope", ha="center",
            va="center", size=11,
            bbox=bbox_props, rotation=-120)
    plt.plot(mac_axes((5.205-h0)+0.25), ax_lims, 'b', zorder=10, lw=5)  # Take-Off Case
    ax.text(mac_axes((5.205-h0)+0.25)[0] - 250, big_weights['OWE_w'] + 1400, "Ground Handling Envelope", ha="center", va="center", size=11,
            bbox=bbox_props, rotation=120)
    plt.plot([mac_axes(0.07)[0], mac_x_point(0.07,big_weights["MZFW"])], [ax_lims[0], big_weights["MZFW"]], 'g', zorder=10, lw=5)  # Landing Case
    plt.plot([mac_x_point(0.07, big_weights["MZFW"]), mac_x_point(0.07, big_weights["MZFW"]) + (fuel_x_delta*0.9)], [big_weights["MZFW"], big_weights["MTOW_w"]], 'g',
             zorder=10, lw=5)
    ax.text(mac_axes(0.07)[0]-50, big_weights['OWE_w'] + 800, "In-Flight Envelope", ha="center", va="center", size=11,
            bbox=bbox_props, rotation=110)

    bbox_props = dict(boxstyle="round", fc="w", ec="0.5", alpha=0.9)
    ax.text(mac_x_point(0.07, big_weights["MZFW"]) + (fuel_x_delta*0.9) - 650, big_weights["MTOW_w"] + 450, "MTOW", ha="center", va="center", size=11,
            bbox=bbox_props)
    ax.text(mac_x_point(0.07, big_weights["MZFW"]) + (fuel_x_delta * 0.9) - 950, big_weights["MZFW"] + 450, "MZFW",
            ha="center", va="center", size=11,
            bbox=bbox_props)

    xtick_hold = []
    for mac_pos in mac_set:
        plt.plot(mac_axes(mac_pos), ax_lims, 'k')
        xtick_hold.append(mac_axes(mac_pos)[0])
        # plt.annotate(mac_pos, [mac_axes(mac_pos)[0], ax_lims[1]-(0.05*(ax_lims[1]-ax_lims[0]))])
    plt.xticks(xtick_hold)

    plt.title("CTA2-100 Weights & Balance Diagram", size=20)

    plt.savefig('vector_plot.png')
    plt.show()


plotit(mac_range=[-0.07, 0.67])
